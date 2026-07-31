from pathlib import Path
from datetime import datetime
from typing import List
import logging

from chatbot_agent import ChatbotAgent
from judge import Judge
from openpyxl import Workbook, load_workbook
import json
import argparse

# GUI 연동
try:
    from chat_gui import ChatGUI
    import tkinter as tk
except Exception:
    ChatGUI = None
    tk = None

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = ROOT / "_OUTPUT" / "results"
LOGS_DIR = ROOT / "_OUTPUT" / "logs"
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_PATH = OUTPUTS_DIR / "chat_history.xlsx"
JSON_PATH = OUTPUTS_DIR / "evaluations.json"
LOG_PATH = LOGS_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"

# Excel 컬럼 순서: 시간/질문/답변 + 8지표 점수 + 평균 + Pass/Fail
from judge import Judge as _Judge  # METRICS 재사용
METRICS = _Judge.METRICS
EXCEL_HEADER = ["시간", "질문", "답변"] + METRICS + ["평균", "Pass/Fail"]


def append_to_excel(path: Path, question: str, answer: str, eval_rec=None) -> None:
    """평가 레코드 전체를 받아 8지표 점수 + 평균 + Pass/Fail을 함께 기록."""
    metrics = (eval_rec or {}).get("metrics", {}) if isinstance(eval_rec, dict) else {}
    overall = (eval_rec or {}).get("overall_score") if isinstance(eval_rec, dict) else None
    passed = (eval_rec or {}).get("pass") if isinstance(eval_rec, dict) else None
    verdict = "" if passed is None else ("PASS" if passed else "FAIL")

    metric_cells = [metrics.get(m, "") if metrics.get(m) is not None else "" for m in METRICS]
    row = (
        [datetime.now().isoformat(), question, answer]
        + metric_cells
        + [overall if overall is not None else "", verdict]
    )

    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADER)
        ws.append(row)
        wb.save(path)
    else:
        wb = load_workbook(filename=path)
        ws = wb.active
        ws.append(row)
        wb.save(path)


def append_to_json(path: Path, eval_rec) -> None:
    """질문/답변/8지표 점수/이유/평균/Pass를 JSON 배열에 누적 저장."""
    if not isinstance(eval_rec, dict):
        return
    data = []
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, list):
                data = []
        except Exception:
            data = []
    data.append(eval_rec)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def write_log(path: Path, lines: List[str]) -> None:
    with open(path, "a", encoding="utf-8") as f:
        for line in lines:
            f.write(line + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('--cli', action='store_true', help='Run in CLI mode')
    args = parser.parse_args()

    agent = ChatbotAgent()
    judge = Judge(agent.client, save_path=OUTPUTS_DIR / "judge_evals.jsonl")

    def save_record(q, a, eval_rec):
        append_to_excel(EXCEL_PATH, q, a, eval_rec)
        append_to_json(JSON_PATH, eval_rec)

    # 로그 헤더 작성
    header = [f"# Chat Session ({datetime.now().isoformat()})", ""]
    write_log(LOG_PATH, header)

    if not args.cli and ChatGUI is not None and tk is not None:
        # GUI 모드 (기본)
        logger.info("GUI 모드로 실행")
        root = tk.Tk()
        app = ChatGUI(root,
                      reply_func=agent.ask,
                      judge=judge,
                      save_callback=save_record,
                      log_callback=lambda lines: write_log(LOG_PATH, lines),
                      clear_callback=lambda: (agent.reset_history(), judge.reset()))
        root.mainloop()
        print(f"엑셀 저장 경로: {EXCEL_PATH}")
        print(f"로그 저장 경로: {LOG_PATH}")
        return

    # 기존 CLI 모드
    logger.info("챗봇 시작 — 종료하려면 '/quit' 입력")
    print("챗봇 시작 — 종료하려면 '/quit' 입력")

    while True:
        try:
            question = input("\n질문: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n종료합니다.")
            break

        if not question:
            continue
        if question.lower() == "/quit":
            print("세션 종료")
            break

        write_log(LOG_PATH, [f"## 질문 ({datetime.now().isoformat()})", "", question, ""])
        print("응답 생성 중...")
        try:
            answer = agent.ask(question)
        except Exception as e:
            err = f"오류: {e}"
            write_log(LOG_PATH, [err, ""])
            print(err)
            continue
        print("\n=== 답변 ===\n")
        print(answer)

        # Judge 평가
        print("\n(평가 중...)")
        try:
            eval_rec = judge.evaluate(question, answer)
        except Exception as e:
            eval_rec = {"error": str(e)}
            logger.warning("Judge 평가 중 오류: %s", e)

        print("\n(저장 중...)")
        append_to_excel(EXCEL_PATH, question, answer, eval_rec)
        append_to_json(JSON_PATH, eval_rec)
        write_log(LOG_PATH, ["### 답변", "", answer, ""])
        # Judge 결과를 로그에 추가
        write_log(LOG_PATH, ["### Judge", "", json.dumps(eval_rec, ensure_ascii=False, indent=2), ""])
        # 세션 누적 평균 / 종합 Pass-Fail
        avg = judge.running_average()
        if avg is not None:
            print(f"누적 평균: {avg:.1f}/{judge.MAX_SCORE}  종합: {judge.session_verdict()}")
        write_log(LOG_PATH, ["---", ""])

    print(f"엑셀 저장 경로: {EXCEL_PATH}")
    print(f"로그 저장 경로: {LOG_PATH}")


if __name__ == "__main__":
    main()
