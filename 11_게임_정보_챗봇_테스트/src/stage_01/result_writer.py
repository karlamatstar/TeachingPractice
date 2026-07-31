
# 게임 가챠 테스트 결과 저장 도우미

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


EXCEL_COLUMNS = [
    "TC_ID",
    "품질지표",
    "검색형태",
    "유형",
    "난이도",
    "질문",
    "답변",
    "기대결과",
    "검증주요요소"
]


def save_json(
    results: list,
    file_path
):
    """
    결과 JSON 저장
    """

    if not results:
        raise ValueError("저장할 결과가 없습니다.")

    file_path = Path(file_path)

    with open(
        file_path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            results,
            f,
            ensure_ascii=False,
            indent=4
        )

    print(
        f"[INFO] JSON 저장 완료 : {file_path}"
    )


def save_excel(
    results: list,
    file_path
):
    """
    결과 Excel 저장
    """

    if not results:
        raise ValueError("저장할 결과가 없습니다.")

    file_path = Path(file_path)

    df = pd.DataFrame(results, columns=EXCEL_COLUMNS)

    # 컬럼 순서 고정
    df = df.reindex(
        columns=EXCEL_COLUMNS
    )

    # Excel 저장
    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Test_Result",
            index=False
        )

    # 스타일 적용
    workbook = load_workbook(
        file_path
    )

    worksheet = workbook["Test_Result"]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    header_font = Font(
        bold=True
    )

    # 헤더 스타일
    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # 자동 너비
    for column in worksheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            try:
                value = str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

            except Exception:
                pass

        adjusted_width = min(
            max(max_length + 2, 15),
            80
        )

        worksheet.column_dimensions[
            column_letter
        ].width = adjusted_width

    # 줄바꿈
    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    worksheet.freeze_panes = "A2"

    workbook.save(file_path)

    print(
        f"[INFO] EXCEL 저장 완료 : {file_path}"
    )


if __name__ == "__main__":

    project_dir = Path(__file__).resolve().parents[2]
    sample_output_dir = project_dir / "_OUTPUT" / "stage_01"
    sample_output_dir.mkdir(parents=True, exist_ok=True)

    sample_data = [
        {
            "TC_ID": "TC001",
            "품질지표": "정확성",
            "검색형태": "키워드",
            "유형": "해피",
            "난이도": "하",
            "질문": "가챠 확률 알려줘",
            "답변": "확률은 3%입니다.",
            "기대결과": "확률 제공",
            "검증주요요소": "확률 수치 포함"
        }
    ]

    save_json(
        sample_data,
        sample_output_dir / "sample.json"
    )

    save_excel(
        sample_data,
        sample_output_dir / "sample.xlsx"
    )
